import os
import re
import shutil
import tempfile
import subprocess
import threading
import time
import webbrowser
import zipfile
from flask import Flask, request, send_file, jsonify, render_template_string
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__)

# Stato globale per il tracciamento dei log e del progresso
app_state = {
    "status": "idle", # idle, processing, done, error
    "progress": 0,
    "logs": [],
    "output_dir": ""
}

def log(msg):
    print(msg)
    app_state["logs"].append(msg)

# ==========================================
# LOGICA DI SPLITTING E FORMATTAZIONE
# ==========================================

def split_italian_name(full_name):
    """
    Divide un nominativo completo (Cognome e Nome) in Cognome e Nome separati.
    Gestisce i prefissi comuni italiani (Di, De, Da, Lo, La, Del, Della, Delle, ecc.).
    """
    if not full_name or not isinstance(full_name, str):
        return "", ""
    
    parts = full_name.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    
    prefixes = {"di", "de", "da", "lo", "la", "li", "del", "della", "delle", "delli", "dell", "van", "von", "d'"}
    
    first_word_lower = parts[0].lower().replace("'", "")
    if (first_word_lower in prefixes or parts[0].endswith("'")) and len(parts) > 2:
        cognome = parts[0] + " " + parts[1]
        nome = " ".join(parts[2:])
    else:
        cognome = parts[0]
        nome = " ".join(parts[1:])
        
    return cognome, nome

def make_safe_filename(cognome, nome):
    """
    Genera un nome file sicuro per il file system.
    """
    clean_cognome = "".join(c for c in cognome if c.isalnum())
    clean_nome = "".join(c for c in nome if c.isalnum())
    if not clean_cognome and not clean_nome:
        return "ricevuta"
    return f"{clean_cognome}{clean_nome}-ricevuta"

def replace_placeholders(text, mapping):
    """
    Sostituisce i segnaposto tipo {{campo}} in modo case-insensitive.
    """
    if not isinstance(text, str):
        return text
    
    def repl(match):
        key = match.group(1).strip().lower()
        if key == 'città':
            key = 'citta'
            
        if key in mapping:
            return str(mapping[key])
            
        col_mappings = {
            'cognome e nome del bambino': mapping.get('nome_bambino_completo', ''),
            'codice fiscale del bambino': mapping.get('cf_bambino', ''),
            'nominativo genitore a cui intestare la ricevuta': mapping.get('nome_genitore_completo', ''),
            'codice fiscale del genitore a cui intestare la ricevuta': mapping.get('cf_genitore', ''),
            'via di residenza': mapping.get('via', ''),
            'citta': mapping.get('citta', ''),
            'città': mapping.get('citta', ''),
            'cap': mapping.get('cap', '')
        }
        if key in col_mappings:
            return str(col_mappings[key])
            
        return match.group(0)

    return re.sub(r'\{\{?([^{}]+)\}\}?', repl, text)

# ==========================================
# GENERAZIONE FILE DI ESEMPIO (DEMO)
# ==========================================

def genera_dati_demo_temporanei():
    """Crea e salva i file demo in una cartella temporanea."""
    temp_dir = tempfile.mkdtemp()
    
    # 1. paganti.xlsx
    excel_dati_path = os.path.join(temp_dir, "paganti.xlsx")
    dati_esempio = {
        "Cognome e Nome del Bambino": ["Rossi Mario", "Bianchi Sofia", "Di Francesco Luca"],
        "Codice Fiscale del Bambino": ["RSSMRA15A01F205G", "BNHSFO16B42H501C", "DFRLCU17C03L219Y"],
        "Nominativo Genitore a cui Intestare la Ricevuta": ["Rossi Giovanni", "Bianchi Beatrice", "Di Francesco Antonio"],
        "Codice Fiscale del Genitore a cui Intestare la Ricevuta": ["RSSGNN80A01F205H", "BNHBRC82B42H501D", "DFRNTN84C03L219Z"],
        "Via di Residenza": ["Via Roma 12", "Via Garibaldi 45", "Corso Cavour 88"],
        "CITTA": ["Perugia", "Milano", "Perugia"],
        "CAP": ["06121", "20121", "06132"]
    }
    pd.DataFrame(dati_esempio).to_excel(excel_dati_path, index=False)
    
    # 2. modello_ricevuta.xlsx
    excel_modello_path = os.path.join(temp_dir, "modello_ricevuta.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ricevuta"
    ws.views.sheetView[0].showGridLines = True
    
    # Stili
    font_titolo = Font(name="Calibri", size=16, bold=True)
    font_sottotitolo = Font(name="Calibri", size=11, bold=True)
    font_sezione = Font(name="Calibri", size=11, bold=True)
    font_etichetta = Font(name="Calibri", size=11)
    font_valore = Font(name="Calibri", size=11, italic=True)
    border_sottile = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )
    
    ws["A1"] = "RICEVUTA"
    ws["A1"].font = font_titolo
    ws["A2"] = "ASSOCIAZIONE TURISTICA PRO PILA - A.P.S."
    ws["A2"].font = font_sottotitolo
    ws["A3"] = "Via del Moggio sc 06132"
    ws["A3"].font = font_etichetta
    ws["A4"] = "Pila - Perugia (PG)"
    ws["A4"].font = font_etichetta
    ws["A5"] = "C.F. 80006780540 - P.IVA 01368930546"
    ws["A5"].font = font_etichetta
    
    # Numero e Data Ricevuta
    ws["E4"] = "N. RICEVUTA"
    ws["E4"].font = font_etichetta
    ws["E4"].border = border_sottile
    ws["F4"] = ""
    ws["F4"].border = border_sottile
    
    ws["E5"] = "DATA"
    ws["E5"].font = font_etichetta
    ws["E5"].border = border_sottile
    ws["F5"] = ""
    ws["F5"].border = border_sottile
    
    # Dati Genitore e Bambino
    ws["A11"] = "DATI DEL GENITORE"
    ws["A11"].font = font_sezione
    ws["E11"] = "DATI DEL BAMBINO"
    ws["E11"].font = font_sezione
    
    # Nome, Cognome, CF
    ws["A13"] = "Nome:"
    ws["A13"].font = font_etichetta
    ws["B13"] = "{{nome_genitore}}"
    ws["B13"].font = font_valore
    ws["E13"] = "Nome:"
    ws["E13"].font = font_etichetta
    ws["F13"] = "{{nome_bambino}}"
    ws["F13"].font = font_valore
    
    ws["A14"] = "Cognome:"
    ws["A14"].font = font_etichetta
    ws["B14"] = "{{cognome_genitore}}"
    ws["B14"].font = font_valore
    ws["E14"] = "Cognome:"
    ws["E14"].font = font_etichetta
    ws["F14"] = "{{cognome_bambino}}"
    ws["F14"].font = font_valore
    
    ws["A15"] = "C.F.:"
    ws["A15"].font = font_etichetta
    ws["B15"] = "{{cf_genitore}}"
    ws["B15"].font = font_valore
    ws["E15"] = "C.F.:"
    ws["E15"].font = font_etichetta
    ws["F15"] = "{{cf_bambino}}"
    ws["F15"].font = font_valore
    
    ws["A18"] = "Residenza:"
    ws["A18"].font = font_sezione
    
    ws["A19"] = "Via"
    ws["A19"].font = font_etichetta
    ws["B19"] = "{{via}}"
    ws["B19"].font = font_valore
    
    ws["A20"] = "cap"
    ws["A20"].font = font_etichetta
    ws["B20"] = "{{CAP}}"
    ws["B20"].font = font_valore
    
    ws["A21"] = "Città"
    ws["A21"].font = font_etichetta
    ws["B21"] = "{{citta}}"
    ws["B21"].font = font_valore
    
    # Tabella
    ws["B24"] = "DESCRIZIONE"
    ws["B24"].font = font_sottotitolo
    ws["B24"].border = border_sottile
    ws["B24"].alignment = Alignment(horizontal="center")
    
    ws["F24"] = "IMPORTO"
    ws["F24"].font = font_sottotitolo
    ws["F24"].border = border_sottile
    ws["F24"].alignment = Alignment(horizontal="center")
    
    settimane = [
        "SETTIMANA DAL 03/08/2026 AL 07/08/2026",
        "SETTIMANA DAL 10/08/2026 AL 14/08/2026",
        "SETTIMANA DAL 17/08/2026 AL 21/08/2026",
        "SETTIMANA DAL 24/08/2026 AL 28/08/2026"
    ]
    
    for idx, sett in enumerate(settimane):
        row_num = 25 + idx
        ws.cell(row=row_num, column=2, value=sett).font = font_etichetta
        ws.cell(row=row_num, column=2).border = border_sottile
        ws.cell(row=row_num, column=6, value=120 if idx == 0 else "").border = border_sottile
        ws.cell(row=row_num, column=6).font = font_etichetta
        
    ws["E30"] = "Totale"
    ws["E30"].font = font_sottotitolo
    ws["E30"].border = border_sottile
    ws["F30"] = "=SUM(F25:F28)"
    ws["F30"].font = font_sottotitolo
    ws["F30"].border = border_sottile
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    
    wb.save(excel_modello_path)
    wb.close()
    
    return excel_dati_path, excel_modello_path, temp_dir

# ==========================================
# ROTTE DI FLASK
# ==========================================

@app.route('/')
def home():
    html_page = """
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Generatore Ricevute PDF</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
        <style>
            :root {
                --bg: #0f172a;
                --card-bg: rgba(30, 41, 59, 0.7);
                --text: #f8fafc;
                --text-muted: #94a3b8;
                --primary: #3b82f6;
                --primary-hover: #2563eb;
                --success: #10b981;
                --success-hover: #059669;
                --border: #334155;
            }
            
            * {
                box-sizing: border-box;
                margin: 0;
                padding: 0;
            }
            
            body {
                font-family: 'Inter', sans-serif;
                background-color: var(--bg);
                color: var(--text);
                min-height: 100vh;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                padding: 20px;
                background-image: radial-gradient(circle at top right, rgba(59, 130, 246, 0.1), transparent 40%),
                                  radial-gradient(circle at bottom left, rgba(16, 185, 129, 0.05), transparent 40%);
            }
            
            .container {
                max-width: 800px;
                width: 100%;
            }
            
            header {
                text-align: center;
                margin-bottom: 30px;
            }
            
            header h1 {
                font-size: 2.2rem;
                font-weight: 700;
                letter-spacing: -0.05em;
                margin-bottom: 8px;
                background: linear-gradient(to right, #3b82f6, #60a5fa, #10b981);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
            }
            
            header p {
                color: var(--text-muted);
                font-size: 1.1rem;
            }
            
            .card {
                background: var(--card-bg);
                backdrop-filter: blur(12px);
                border: 1px solid var(--border);
                border-radius: 16px;
                padding: 30px;
                box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.3), 0 8px 10px -6px rgba(0, 0, 0, 0.3);
                margin-bottom: 25px;
            }
            
            .grid-inputs {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin-bottom: 25px;
            }
            
            @media (max-width: 600px) {
                .grid-inputs {
                    grid-template-columns: 1fr;
                }
            }
            
            .upload-box {
                border: 2px dashed var(--border);
                border-radius: 12px;
                padding: 25px 15px;
                text-align: center;
                cursor: pointer;
                transition: all 0.2s ease;
                background: rgba(15, 23, 42, 0.3);
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                min-height: 160px;
            }
            
            .upload-box:hover {
                border-color: var(--primary);
                background: rgba(59, 130, 246, 0.05);
            }
            
            .upload-box.dragover {
                border-color: var(--success);
                background: rgba(16, 185, 129, 0.1);
            }
            
            .upload-box svg {
                width: 40px;
                height: 40px;
                stroke: var(--text-muted);
                margin-bottom: 12px;
                transition: stroke 0.2s ease;
            }
            
            .upload-box:hover svg {
                stroke: var(--primary);
            }
            
            .upload-box p {
                font-size: 0.9rem;
                font-weight: 500;
                margin-bottom: 4px;
            }
            
            .upload-box span {
                font-size: 0.75rem;
                color: var(--text-muted);
            }
            
            .file-selected {
                color: var(--success) !important;
                font-weight: 600 !important;
            }
            
            .btn {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 100%;
                padding: 14px 24px;
                border-radius: 8px;
                font-weight: 600;
                font-size: 1rem;
                border: none;
                cursor: pointer;
                transition: all 0.2s ease;
                gap: 8px;
            }
            
            .btn-primary {
                background: var(--primary);
                color: white;
                box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3);
            }
            
            .btn-primary:hover:not(:disabled) {
                background: var(--primary-hover);
                transform: translateY(-1px);
            }
            
            .btn-primary:disabled {
                background: #1e293b;
                color: var(--text-muted);
                cursor: not-allowed;
                border: 1px solid var(--border);
                box-shadow: none;
            }
            
            .btn-demo {
                background: transparent;
                border: 1px solid var(--success);
                color: var(--success);
                margin-bottom: 20px;
            }
            
            .btn-demo:hover {
                background: rgba(16, 185, 129, 0.1);
            }
            
            /* Console log styling */
            .console-box {
                background: #020617;
                border: 1px solid var(--border);
                border-radius: 8px;
                padding: 15px;
                font-family: 'Consolas', monospace;
                font-size: 0.85rem;
                height: 150px;
                overflow-y: auto;
                color: #38bdf8;
                margin-top: 20px;
                display: none;
            }
            
            .progress-container {
                margin-top: 20px;
                display: none;
            }
            
            .progress-text {
                display: flex;
                justify-content: space-between;
                font-size: 0.85rem;
                color: var(--text-muted);
                margin-bottom: 6px;
            }
            
            .progress-bar {
                height: 8px;
                width: 100%;
                background: var(--border);
                border-radius: 4px;
                overflow: hidden;
            }
            
            .progress-fill {
                height: 100%;
                width: 0%;
                background: linear-gradient(to right, var(--primary), var(--success));
                transition: width 0.3s ease;
            }
            
            /* Placeholders table */
            .guide-title {
                font-size: 1.1rem;
                font-weight: 600;
                margin-bottom: 12px;
                display: flex;
                align-items: center;
                gap: 6px;
            }
            
            .guide-list {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 10px;
                font-size: 0.85rem;
                color: var(--text-muted);
                background: rgba(15, 23, 42, 0.4);
                padding: 15px;
                border-radius: 8px;
                border: 1px solid var(--border);
            }
            
            .guide-item strong {
                color: var(--text);
                font-family: monospace;
                background: rgba(255, 255, 255, 0.05);
                padding: 2px 4px;
                border-radius: 4px;
            }
            
            footer {
                text-align: center;
                margin-top: 20px;
                color: var(--text-muted);
                font-size: 0.8rem;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <header>
                <h1>Generatore Ricevute PDF</h1>
                <p>Inserisci i campi sul tuo foglio Excel e crea i PDF in un clic</p>
            </header>
            
            <div style="text-align: center;">
                <button class="btn btn-demo" onclick="scaricaDemo()">
                    📥 Scarica Pacchetto Demo (Dati + Modello Esempio)
                </button>
            </div>
            
            <div class="card">
                <form id="uploadForm">
                    <div class="grid-inputs">
                        <!-- Excel Dati -->
                        <div class="upload-box" id="dropDati" onclick="document.getElementById('inputDati').click()">
                            <svg fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 17v-2m3 2v-4m3 4v-6m2 10H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"></path>
                            </svg>
                            <p id="labelDati">Excel dei Paganti (Dati)</p>
                            <span id="subDati">Trascina qui o clicca per caricare</span>
                            <input type="file" id="inputDati" name="excel_data" accept=".xlsx,.xls,.xlxsx" style="display:none;" onchange="handleFileSelected(this, 'Dati')">
                        </div>
                        
                        <!-- Excel Modello -->
                        <div class="upload-box" id="dropModello" onclick="document.getElementById('inputModello').click()">
                            <svg fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M4 16l4.586-4.586a2 2 0 012.828 0L16 16m-2-2l1.586-1.586a2 2 0 012.828 0L20 14m-6-6h.01M6 20h12a2 2 0 002-2V6a2 2 0 00-2-2H6a2 2 0 00-2 2v12a2 2 0 002 2z"></path>
                            </svg>
                            <p id="labelModello">Modello Ricevuta (Excel)</p>
                            <span id="subModello">Trascina qui o clicca per caricare</span>
                            <input type="file" id="inputModello" name="excel_template" accept=".xlsx,.xls" style="display:none;" onchange="handleFileSelected(this, 'Modello')">
                        </div>
                    </div>
                    
                    <button type="button" id="btnGenera" class="btn btn-primary" onclick="avviaGenerazione()" disabled>
                        ⚡ Compila e Genera PDF (.zip)
                    </button>
                </form>
                
                <div class="progress-container" id="progressContainer">
                    <div class="progress-text">
                        <span id="statusText">Inizializzazione...</span>
                        <span id="progressPercent">0%</span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" id="progressFill"></div>
                    </div>
                </div>
                
                <div class="console-box" id="consoleBox"></div>
            </div>
            
            <div class="card">
                <div class="guide-title">
                    📋 Guida ai Segnaposto da inserire nel file Modello Excel
                </div>
                <div class="guide-list">
                    <div class="guide-item">👦 <strong>{{nome_bambino}}</strong> : Nome Bambino</div>
                    <div class="guide-item">👦 <strong>{{cognome_bambino}}</strong> : Cognome Bambino</div>
                    <div class="guide-item">👦 <strong>{{nome_bambino_completo}}</strong> : Nome intero</div>
                    <div class="guide-item">👦 <strong>{{cf_bambino}}</strong> : CF del Bambino</div>
                    <div class="guide-item">👨 <strong>{{nome_genitore}}</strong> : Nome Genitore</div>
                    <div class="guide-item">👨 <strong>{{cognome_genitore}}</strong> : Cognome Genitore</div>
                    <div class="guide-item">👨 <strong>{{nome_genitore_completo}}</strong> : Nome intero</div>
                    <div class="guide-item">👨 <strong>{{cf_genitore}}</strong> : CF del Genitore</div>
                    <div class="guide-item">📍 <strong>{{via}}</strong> : Via di Residenza</div>
                    <div class="guide-item">📍 <strong>{{citta}}</strong> : Città di Residenza</div>
                    <div class="guide-item">📍 <strong>{{cap}}</strong> / <strong>{{CAP}}</strong> : CAP</div>
                </div>
            </div>
            
            <footer>
                Associazione Turistica Pro Pila • Sviluppato per la gestione ricevute
            </footer>
        </div>

        <script>
            // Gestione Drag & Drop
            function setupDragAndDrop(dropAreaId, inputId, labelId, type) {
                const dropArea = document.getElementById(dropAreaId);
                const input = document.getElementById(inputId);
                
                ['dragenter', 'dragover'].forEach(eventName => {
                    dropArea.addEventListener(eventName, (e) => {
                        e.preventDefault();
                        dropArea.classList.add('dragover');
                    }, false);
                });
                
                ['dragleave', 'drop'].forEach(eventName => {
                    dropArea.addEventListener(eventName, (e) => {
                        e.preventDefault();
                        dropArea.classList.remove('dragover');
                    }, false);
                });
                
                dropArea.addEventListener('drop', (e) => {
                    const dt = e.dataTransfer;
                    const files = dt.files;
                    if(files.length > 0) {
                        input.files = files;
                        handleFileSelected(input, type);
                    }
                }, false);
            }
            
            setupDragAndDrop('dropDati', 'inputDati', 'labelDati', 'Dati');
            setupDragAndDrop('dropModello', 'inputModello', 'labelModello', 'Modello');
            
            function handleFileSelected(input, type) {
                const label = document.getElementById('label' + type);
                const sub = document.getElementById('sub' + type);
                if (input.files.length > 0) {
                    label.innerText = input.files[0].name;
                    label.classList.add('file-selected');
                    sub.innerText = "File pronto";
                } else {
                    label.innerText = type === 'Dati' ? "Excel dei Paganti (Dati)" : "Modello Ricevuta (Excel)";
                    label.classList.remove('file-selected');
                    sub.innerText = "Trascina qui o clicca per caricare";
                }
                
                // Abilita bottone se entrambi i file sono inseriti
                const inputDati = document.getElementById('inputDati');
                const inputModello = document.getElementById('inputModello');
                document.getElementById('btnGenera').disabled = !(inputDati.files.length > 0 && inputModello.files.length > 0);
            }
            
            function scaricaDemo() {
                window.location.href = "/demo-download";
            }
            
            let pollInterval;
            
            function avviaGenerazione() {
                const formData = new FormData(document.getElementById('uploadForm'));
                
                // Mostra elementi grafici di stato
                document.getElementById('progressContainer').style.display = 'block';
                document.getElementById('consoleBox').style.display = 'block';
                document.getElementById('btnGenera').disabled = true;
                
                // Reset console
                const consoleBox = document.getElementById('consoleBox');
                consoleBox.innerHTML = 'Avvio del caricamento dei file...<br>';
                
                // Avvia polling dei log
                pollInterval = setInterval(aggiornaStato, 800);
                
                fetch('/generate', {
                    method: 'POST',
                    body: formData
                })
                .then(response => {
                    if (response.ok) {
                        return response.blob();
                    } else {
                        return response.json().then(err => { throw new Error(err.error || 'Errore sconosciuto') });
                    }
                })
                .then(blob => {
                    clearInterval(pollInterval);
                    aggiornaStato(); // un'ultima lettura dello stato
                    
                    // Crea link per scaricare il file zip generato
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.style.display = 'none';
                    a.href = url;
                    a.download = 'ricevute_pdf.zip';
                    document.body.appendChild(a);
                    a.click();
                    window.URL.revokeObjectURL(url);
                    
                    document.getElementById('statusText').innerText = "Elaborazione completata!";
                    document.getElementById('btnGenera').disabled = false;
                })
                .catch(error => {
                    clearInterval(pollInterval);
                    consoleBox.innerHTML += `<br><span style="color: #ef4444;">ERRORE: ${error.message}</span>`;
                    document.getElementById('statusText').innerText = "Errore durante l'elaborazione.";
                    document.getElementById('btnGenera').disabled = false;
                });
            }
            
            function aggiornaStato() {
                fetch('/status')
                .then(res => res.json())
                .then(data => {
                    document.getElementById('progressPercent').innerText = data.progress + '%';
                    document.getElementById('progressFill').style.width = data.progress + '%';
                    document.getElementById('statusText').innerText = data.status === 'processing' ? 'Elaborazione in corso...' : 'In attesa...';
                    
                    const consoleBox = document.getElementById('consoleBox');
                    consoleBox.innerHTML = data.logs.join('<br>');
                    consoleBox.scrollTop = consoleBox.scrollHeight;
                });
            }
        </script>
    </body>
    </html>
    """
    return render_template_string(html_page)

@app.route('/demo-download')
def demo_download():
    # Genera i file demo al volo
    dati, modello, temp_dir = genera_dati_demo_temporanei()
    
    # Comprime in uno zip
    zip_path = os.path.join(temp_dir, "demo_ricevute.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        zipf.write(dati, arcname="paganti.xlsx")
        zipf.write(modello, arcname="modello_ricevuta.xlsx")
        
    return send_file(zip_path, as_attachment=True, download_name="demo_ricevute.zip")

@app.route('/status')
def get_status():
    return jsonify(app_state)

@app.route('/generate', methods=['POST'])
def generate():
    if 'excel_data' not in request.files or 'excel_template' not in request.files:
        return jsonify({"error": "Entrambi i file Excel sono richiesti"}), 400
        
    file_data = request.files['excel_data']
    file_template = request.files['excel_template']
    
    if file_data.filename == '' or file_template.filename == '':
        return jsonify({"error": "Nessun file selezionato"}), 400
        
    # Crea cartella temporanea per elaborazione
    temp_work_dir = tempfile.mkdtemp()
    
    dati_path = os.path.join(temp_work_dir, "dati.xlsx")
    template_path = os.path.join(temp_work_dir, "modello.xlsx")
    
    file_data.save(dati_path)
    file_template.save(template_path)
    
    # Cartelle di output
    output_pdf_temp = os.path.join(temp_work_dir, "pdf_generati")
    os.makedirs(output_pdf_temp, exist_ok=True)
    
    # Cartella di output fissa nel workspace per comodità
    workspace_output = os.path.abspath(os.path.join(os.getcwd(), "output_ricevute"))
    os.makedirs(workspace_output, exist_ok=True)
    
    # Resetta stato
    app_state["status"] = "processing"
    app_state["progress"] = 0
    app_state["logs"] = []
    
    def process_thread():
        try:
            # Esegui la conversione
            log("Inizio elaborazione ricevute...")
            
            df = pd.read_excel(dati_path, dtype=str)
            df.columns = [c.strip() for c in df.columns]
            
            # Controllo colonne
            required_cols = [
                "Cognome e Nome del Bambino",
                "Codice Fiscale del Bambino",
                "Nominativo Genitore a cui Intestare la Ricevuta",
                "Codice Fiscale del Genitore a cui Intestare la Ricevuta",
                "Via di Residenza",
                "CITTA",
                "CAP"
            ]
            missing_cols = [c for c in required_cols if c not in df.columns]
            if missing_cols:
                raise Exception(f"Colonne mancanti nell'Excel dei dati: {', '.join(missing_cols)}")
                
            total_rows = len(df)
            log(f"Totale righe trovate da elaborare: {total_rows}")
            
            def safe_str(val):
                if pd.isna(val):
                    return ""
                s = str(val).strip()
                if s.endswith(".0"):
                    s = s[:-2]
                if s.lower() == 'nan':
                    return ""
                return s

            for index, row in df.iterrows():
                nominativo_bambino = safe_str(row["Cognome e Nome del Bambino"])
                cf_bambino = safe_str(row["Codice Fiscale del Bambino"])
                nominativo_genitore = safe_str(row["Nominativo Genitore a cui Intestare la Ricevuta"])
                cf_genitore = safe_str(row["Codice Fiscale del Genitore a cui Intestare la Ricevuta"])
                via = safe_str(row["Via di Residenza"])
                citta = safe_str(row["CITTA"])
                cap = safe_str(row["CAP"])

                
                cognome_bambino, nome_bambino = split_italian_name(nominativo_bambino)
                cognome_genitore, nome_genitore = split_italian_name(nominativo_genitore)
                
                row_data = {
                    'nome_bambino_completo': nominativo_bambino,
                    'cognome_bambino': cognome_bambino,
                    'nome_bambino': nome_bambino,
                    'cf_bambino': cf_bambino,
                    'nome_genitore_completo': nominativo_genitore,
                    'cognome_genitore': cognome_genitore,
                    'nome_genitore': nome_genitore,
                    'cf_genitore': cf_genitore,
                    'via': via,
                    'citta': citta,
                    'cap': cap
                }
                
                filename_base = make_safe_filename(cognome_bambino, nome_bambino)
                generated_pdf_name = f"{filename_base}.pdf"
                dest_pdf = os.path.join(workspace_output, generated_pdf_name)
                
                if os.path.exists(dest_pdf):
                    log(f"[{index+1}/{total_rows}] Ricevuta già esistente per {nominativo_bambino} ({generated_pdf_name}). Salto.")
                    shutil.copy2(dest_pdf, os.path.join(output_pdf_temp, generated_pdf_name))
                    app_state["progress"] = int(((index + 1) / total_rows) * 100)
                    continue
                
                # Compila Excel
                wb = openpyxl.load_workbook(template_path)

                for ws in wb.worksheets:
                    for r in ws.iter_rows():
                        for cell in r:
                            if cell.value is not None and isinstance(cell.value, str):
                                cell.value = replace_placeholders(cell.value, row_data)
                
                temp_xlsx = os.path.join(temp_work_dir, f"{filename_base}.xlsx")
                wb.save(temp_xlsx)
                wb.close()

                
                log(f"[{index+1}/{total_rows}] Compilato Excel temporaneo per {nominativo_bambino}")
                
                # Conversione in PDF usando LibreOffice
                cmd = [
                    "libreoffice",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    output_pdf_temp,
                    temp_xlsx
                ]
                
                result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                if result.returncode != 0:
                    # Alternativa soffice
                    cmd[0] = "soffice"
                    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    if result.returncode != 0:
                        raise Exception(f"Conversione LibreOffice fallita per {nominativo_bambino}: {result.stderr}")
                
                # Copia anche il PDF risultante nella cartella del Workspace per persistenza
                generated_pdf_name = f"{filename_base}.pdf"
                src_pdf = os.path.join(output_pdf_temp, generated_pdf_name)
                dest_pdf = os.path.join(workspace_output, generated_pdf_name)
                shutil.copy2(src_pdf, dest_pdf)
                
                # Rimuove Excel temporaneo
                if os.path.exists(temp_xlsx):
                    os.remove(temp_xlsx)
                    
                log(f"[{index+1}/{total_rows}] PDF generato e salvato con successo: {generated_pdf_name}")
                app_state["progress"] = int(((index + 1) / total_rows) * 100)
                
            log("Tutti i file PDF sono stati creati con successo!")
            log(f"I PDF sono stati salvati anche sul disco locale in: {workspace_output}")
            app_state["status"] = "done"
            
        except Exception as e:
            log(f"ERRORE DURANTE L'ELABORAZIONE: {str(e)}")
            app_state["status"] = "error"
            
    # Avvia l'elaborazione in un thread separato per consentire alla richiesta HTTP di rispondere
    t = threading.Thread(target=process_thread)
    t.start()
    
    # Aspetta che il thread completi (o vada in errore)
    while app_state["status"] == "processing":
        time.sleep(0.5)
        
    if app_state["status"] == "error":
        return jsonify({"error": app_state["logs"][-1]}), 500
        
    # Crea ZIP dei file generati per inviarli al browser
    zip_path = os.path.join(temp_work_dir, "ricevute_pdf.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for f in os.listdir(output_pdf_temp):
            if f.endswith('.pdf'):
                zipf.write(os.path.join(output_pdf_temp, f), arcname=f)
                
    # Invia lo zip
    return send_file(zip_path, as_attachment=True, download_name="ricevute_pdf.zip")

# ==========================================
# AVVIO BROWSER AUTOMATICO
# ==========================================

def open_browser():
    # Aspetta un secondo che Flask si avvii
    time.sleep(1.5)
    url = "http://127.0.0.1:5000"
    log(f"Apertura automatica del browser all'indirizzo: {url}")
    webbrowser.open(url)

if __name__ == "__main__":
    # Avvia thread per aprire il browser
    threading.Thread(target=open_browser, daemon=True).start()
    
    # Avvia il server Flask
    app.run(host="127.0.0.1", port=5000, debug=False)
